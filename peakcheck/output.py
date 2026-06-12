"""
Output writers: Excel (.xlsx) and a single ASCII .csv with the same content.

Both formats share the same five sections — Intensities, Positions, Presence,
Fit_Data, Components_Raw — plus a Parameters block that documents the run
(reference path, every component path, transform, profile, χ²/R² and the
numpy/scipy/Python versions). The Excel writes one sheet per section; the
CSV writes one file with labelled `# === SECTION: ... ===` markers, so it
stays human-readable and Origin-importable.
"""
from __future__ import annotations

import collections
import csv
import datetime
import os
import sys

import numpy as np
import scipy

from .background import subtract_background_cfg
from .fit import compute_model
from .io import load_xy
from .plots import _outpath
from . import __version__   # used in the Parameters block (gather_parameters)


def gather_parameters(cfg, peak_centers, picker_settings, stats, had_errors,
                      component_files=None, width_scales=None, fwhms=None):
    """Collect all run parameters + fit statistics into an ordered dict.

    Parameters
    ----------
    component_files : list of str, optional
        Absolute paths of the component data sets that were analysed; listed in
        the output for full provenance.
    width_scales : numpy.ndarray, optional
        Per-peak width scale factors from the optional width refinement (``None``
        when widths were held fixed).
    fwhms : numpy.ndarray, optional
        Per-peak effective FWHM (recorded for provenance).
    """
    unit = f" [{cfg.x_unit}]" if cfg.x_unit else ""
    p = collections.OrderedDict()
    p["Program"]            = "PeakCheck"
    p["Version"]            = __version__
    p["Timestamp"]          = datetime.datetime.now().isoformat(timespec="seconds")
    p["Reference file"]     = os.path.abspath(cfg.reference_file)
    if component_files:
        p["N component files"] = len(component_files)
        # Listed individually so every input path is captured in the output.
        for i, cf in enumerate(component_files, 1):
            p[f"Component {i}"] = os.path.abspath(cf)
    else:
        p["N component files"] = 0
    p["x label"]            = cfg.x_label
    p["y label"]            = cfg.y_label
    p["x unit"]             = cfg.x_unit or "(none)"
    p["x transform"]        = f"x*{cfg.x_scale} + {cfg.x_offset}"
    p["Errors from file"]   = "yes" if had_errors else "no (estimated from data)"
    p["Profile"]            = str(cfg.profile).lower()
    p[f"WG (Gauss FWHM){unit}"]   = cfg.wg
    p[f"WL (Lorentz FWHM){unit}"] = cfg.wl
    p[f"sigma{unit}"]       = round(cfg.sigma(), 6)
    p[f"gamma{unit}"]       = round(cfg.gamma(), 6)
    p[f"Effective FWHM{unit}"] = round(cfg.effective_fwhm(), 4)
    # width refinement provenance
    if width_scales is not None:
        import numpy as _np
        p["Width refinement"] = f"on (mode '{cfg.width_mode}')"
        p["Width bounds [x instrumental]"] = (
            f"{cfg.width_min_factor:g} .. {cfg.width_max_factor:g}")
        if fwhms is not None:
            p[f"Per-peak FWHM{unit}"] = ", ".join(
                f"{f:.3f}" for f in _np.atleast_1d(fwhms))
            p["Per-peak width factor"] = ", ".join(
                f"{s:.3f}" for s in _np.atleast_1d(width_scales))
    else:
        p["Width refinement"] = "off (fixed instrumental width)"
    p[f"x_min{unit}"]       = cfg.x_min
    p[f"x_max{unit}"]       = cfg.x_max
    p["BG window [pts]"]    = cfg.bg_window
    p[f"Tolerance{unit}"]   = cfg.tolerance
    p["SNR threshold"]      = cfg.snr_thresh
    p["Presence SNR on"]    = ("background-corrected signal"
                               if cfg.presence_baseline_corrected else "raw signal")
    if picker_settings:
        p["Prominence (final)"]         = round(picker_settings.get("prominence", float("nan")), 4)
        p["Min distance (final) [pts]"] = picker_settings.get("distance")
        p["Min SNR (final)"]            = picker_settings.get("min_snr")
    p["N peaks"]            = len(peak_centers)
    p[f"Peak centers{unit}"] = ", ".join(f"{c:.2f}" for c in peak_centers)
    if stats is not None:
        p["chi2"]           = round(stats["chi2"], 4)
        p["chi2_reduced"]   = round(stats["chi2_red"], 4)
        p["R2"]             = round(stats["r2"], 6)
        p["DOF"]            = stats["dof"]
        p["N active peaks"] = stats["n_active"]
        p["N fit points"]   = stats["n_points"]
        if stats.get("error_mode") == "scaled":
            p["Std_Err definition"] = (
                f"statistical x sqrt(chi2_red) = x{stats.get('error_scale', 1.0):.4g}")
        else:
            p["Std_Err definition"] = (
                "statistical 1-sigma from (A^T W^2 A)^-1, unscaled")
    p["numpy"]  = np.__version__
    p["scipy"]  = scipy.__version__
    p["Python"] = sys.version.split()[0]
    return p


# ============================================================
#  EXCEL OUTPUT
# ============================================================

def save_excel(cfg, peak_centers, amplitudes, component_results,
               fit_data=None, stats=None, params=None, sigmas=None, gammas=None):
    """Excel workbook with Intensities, Positions, Presence, Fit_Data, Parameters."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
    HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    UNIT_FILL     = PatternFill("solid", start_color="D6E4F0")
    UNIT_FONT     = Font(name="Arial", italic=True, color="1F4E79", size=9)
    DATA_FONT     = Font(name="Arial", size=9)
    FOUND_FILL    = PatternFill("solid", start_color="C6EFCE")
    NOTFOUND_FILL = PatternFill("solid", start_color="FFC7CE")
    FOUND_FONT    = Font(name="Arial", size=9, color="276221")
    NOTFOUND_FONT = Font(name="Arial", size=9, color="9C0006")
    THIN   = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    xu = cfg.x_unit if cfg.x_unit else "x"
    comp_labels = [os.path.splitext(os.path.basename(f))[0]
                   for f, _, _, _ in component_results]

    def write_header(ws, headers, units):
        for col, (h, u) in enumerate(zip(headers, units), start=1):
            hc = ws.cell(row=1, column=col, value=h)
            hc.font, hc.fill, hc.border = HEADER_FONT, HEADER_FILL, BORDER
            hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            uc = ws.cell(row=2, column=col, value=u)
            uc.font, uc.fill, uc.border = UNIT_FONT, UNIT_FILL, BORDER
            uc.alignment = Alignment(horizontal="center")

    wb = Workbook()

    # Sheet 1: Intensities
    ws1 = wb.active
    ws1.title = "Intensities"
    stderr = stats["stderr"] if stats is not None else [float("nan")] * len(peak_centers)
    write_header(ws1,
                 [f"Peak [{xu}]", "Amplitude", "Std_Err"] + comp_labels,
                 [xu, "area", "area"] + ["intensity"] * len(comp_labels))
    for i, (center, amp) in enumerate(zip(peak_centers, amplitudes), start=3):
        ws1.cell(row=i, column=1, value=round(float(center), 4)).border = BORDER
        ws1.cell(row=i, column=2, value=round(float(amp), 4)).border = BORDER
        se = stderr[i - 3]
        sc = ws1.cell(row=i, column=3); sc.border, sc.font = BORDER, DATA_FONT
        if se is not None and np.isfinite(se):
            sc.value = round(float(se), 4)
        for j, (_, fmask, fint, _) in enumerate(component_results, start=4):
            c = ws1.cell(row=i, column=j); c.border, c.font = BORDER, DATA_FONT
            idx = i - 3
            if fmask[idx] and not np.isnan(fint[idx]):
                c.value = round(float(fint[idx]), 4)
    for col, w in (("A", 13), ("B", 13), ("C", 13)):
        ws1.column_dimensions[col].width = w
    for col in range(4, 4 + len(comp_labels)):
        ws1.column_dimensions[get_column_letter(col)].width = 12
    ws1.row_dimensions[1].height = 30
    ws1.freeze_panes = "D3"

    # Sheet 2: Positions
    ws2 = wb.create_sheet("Positions")
    write_header(ws2, [f"Peak_ref [{xu}]"] + comp_labels, [xu] + [xu] * len(comp_labels))
    for i, center in enumerate(peak_centers, start=3):
        ws2.cell(row=i, column=1, value=round(float(center), 4)).border = BORDER
        idx = i - 3
        for j, (_, fmask, _, fpos) in enumerate(component_results, start=2):
            c = ws2.cell(row=i, column=j); c.border, c.font = BORDER, DATA_FONT
            if fmask[idx] and not np.isnan(fpos[idx]):
                c.value = round(float(fpos[idx]), 4)
    ws2.column_dimensions["A"].width = 15
    for col in range(2, 2 + len(comp_labels)):
        ws2.column_dimensions[get_column_letter(col)].width = 12
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = "B3"

    # Sheet 3: Presence
    ws3 = wb.create_sheet("Presence")
    write_header(ws3, [f"Peak [{xu}]"] + comp_labels, [xu] + [""] * len(comp_labels))
    for i, center in enumerate(peak_centers, start=3):
        ws3.cell(row=i, column=1, value=round(float(center), 4)).border = BORDER
        idx = i - 3
        for j, (_, fmask, _, _) in enumerate(component_results, start=2):
            c = ws3.cell(row=i, column=j); c.border = BORDER
            c.alignment = Alignment(horizontal="center")
            if fmask[idx]:
                c.value, c.fill, c.font = 1, FOUND_FILL, FOUND_FONT
            else:
                c.value, c.fill, c.font = 0, NOTFOUND_FILL, NOTFOUND_FONT
    ws3.column_dimensions["A"].width = 13
    for col in range(2, 2 + len(comp_labels)):
        ws3.column_dimensions[get_column_letter(col)].width = 10
    ws3.row_dimensions[1].height = 30
    ws3.freeze_panes = "B3"

    # Sheet 4: Fit_Data
    if fit_data is not None:
        x_f, y_f, yerr_f, amps_f, centers_f = fit_data
        y_corr_f = subtract_background_cfg(y_f, cfg)
        _sig = cfg.sigma() if sigmas is None else sigmas
        _gam = cfg.gamma() if gammas is None else gammas
        total_f, comps_f = compute_model(x_f, amps_f, centers_f,
                                         _sig, _gam, cfg.profile)
        ws4 = wb.create_sheet("Fit_Data")
        peak_headers = [f"{c:.1f} {xu}" for c in centers_f]
        headers4 = [cfg.x_label, "Raw_y", "Error", "BG_corrected", "Sum_Fit"] + peak_headers
        units4   = [xu, cfg.y_label, cfg.y_label, cfg.y_label, cfg.y_label] + [cfg.y_label] * len(centers_f)
        write_header(ws4, headers4, units4)
        for row_i, xi in enumerate(x_f, start=3):
            j = row_i - 3
            row_data = [float(xi), float(y_f[j]), float(yerr_f[j]),
                        float(y_corr_f[j]), float(total_f[j])] + [float(c[j]) for c in comps_f]
            for col_i, val in enumerate(row_data, start=1):
                cell = ws4.cell(row=row_i, column=col_i, value=round(val, 6))
                cell.border, cell.font = BORDER, DATA_FONT
        for col_i in range(1, len(headers4) + 1):
            ws4.column_dimensions[get_column_letter(col_i)].width = 13
        ws4.row_dimensions[1].height = 30
        ws4.freeze_panes = "B3"

    # Sheet 5 (optional): Components_Raw — every component's raw (x, y, error)
    # stored as side-by-side column blocks so each component keeps its own grid.
    # Useful for replotting "reference vs components" in Origin or similar.
    if component_results:
        ws_cr = wb.create_sheet("Components_Raw")
        # Per-component block: 3 columns (x, y, error) — sigma/gamma estimate when no error
        block_cols = 3
        headers_cr, units_cr = [], []
        max_len = 0
        for cf, _fmask, _fint, _fpos in component_results:
            try:
                xs, ys, es, _had = load_xy(cf, cfg)
                max_len = max(max_len, len(xs))
            except Exception:
                xs = ys = es = None
            label = os.path.splitext(os.path.basename(cf))[0]
            headers_cr.extend([f"{label} : {cfg.x_label}",
                               f"{label} : {cfg.y_label}",
                               f"{label} : error"])
            units_cr.extend([xu, cfg.y_label, cfg.y_label])
        write_header(ws_cr, headers_cr, units_cr)
        # Write block by block (re-read once more to keep memory low)
        col_offset = 1
        for cf, _fm, _fi, _fp in component_results:
            try:
                xs, ys, es, _ = load_xy(cf, cfg)
            except Exception:
                col_offset += block_cols
                continue
            for k in range(len(xs)):
                row = k + 3
                ws_cr.cell(row=row, column=col_offset    , value=round(float(xs[k]), 6))
                ws_cr.cell(row=row, column=col_offset + 1, value=round(float(ys[k]), 6))
                ws_cr.cell(row=row, column=col_offset + 2, value=round(float(es[k]), 6))
                for dc in range(block_cols):
                    c = ws_cr.cell(row=row, column=col_offset + dc)
                    c.border, c.font = BORDER, DATA_FONT
            col_offset += block_cols
        for col_i in range(1, len(headers_cr) + 1):
            ws_cr.column_dimensions[get_column_letter(col_i)].width = 14
        ws_cr.row_dimensions[1].height = 30
        ws_cr.freeze_panes = "A3"

    # Sheet 6: Parameters
    if params is not None:
        from openpyxl.styles import Font as _Font
        ws5 = wb.create_sheet("Parameters")
        write_header(ws5, ["Parameter", "Value"], ["", ""])
        for r, (key, val) in enumerate(params.items(), start=3):
            kc = ws5.cell(row=r, column=1, value=str(key))
            vc = ws5.cell(row=r, column=2, value="" if val is None else str(val))
            kc.border, kc.font = BORDER, _Font(name="Arial", size=9, bold=True)
            vc.border, vc.font = BORDER, DATA_FONT
        ws5.column_dimensions["A"].width = 32
        ws5.column_dimensions["B"].width = 60
        ws5.row_dimensions[1].height = 30
        ws5.freeze_panes = "A3"

    outname = _outpath(cfg, cfg.reference_file, "_results.xlsx")
    wb.save(outname)
    print(f"\n  Excel saved: {outname}")
    return outname


# ============================================================
#  CSV OUTPUT  (plain 7-bit ASCII)
# ============================================================

def _ascii(s):
    """Force a string to 7-bit ASCII, replacing anything else with '?'."""
    return str(s).encode("ascii", "replace").decode("ascii")


def _csv_header_lines(params):
    lines = ["# PeakCheck results"]
    if params:
        for key, val in params.items():
            lines.append(_ascii(f"#   {key}: {'' if val is None else val}"))
    return lines


def save_csv(cfg, peak_centers, amplitudes, component_results,
             fit_data=None, stats=None, params=None, sigmas=None, gammas=None):
    """Plain 7-bit-ASCII CSV with all results in a single file.

    The file starts with the full parameter block as commented `#` lines, then
    five clearly labelled sections separated by `# === SECTION: <name> ===`
    headers:

        Intensities
        Positions
        Presence
        Fit_Data            (only if `fit_data` is given)
        Components_Raw      (only if any components were analysed)

    Each section is a normal CSV table with its own column header row, so the
    file is also importable as a single table by readers that ignore blank
    lines (Origin and most others)."""
    out_path    = _outpath(cfg, cfg.reference_file, "_results.csv")
    comp_labels = [_ascii(os.path.splitext(os.path.basename(f))[0])
                   for f, _, _, _ in component_results]
    stderr      = stats["stderr"] if stats is not None else [float("nan")] * len(peak_centers)

    def _fmt(v):
        return "" if (v is None or (isinstance(v, float) and np.isnan(v))) else f"{v:.6g}"

    with open(out_path, "w", newline="", encoding="ascii", errors="replace") as fh:
        # 1) Parameter block (once, at the top)
        for ln in _csv_header_lines(params):
            fh.write(ln + "\n")
        fh.write("#\n")
        w = csv.writer(fh)

        def section(name):
            fh.write("\n")
            fh.write(f"# === SECTION: {name} ===\n")

        # 2) Intensities
        section("Intensities")
        w.writerow(["Peak", "Amplitude", "Std_Err"] + comp_labels)
        for idx, (center, amp) in enumerate(zip(peak_centers, amplitudes)):
            row = [f"{center:.4f}", f"{amp:.4f}", _fmt(float(stderr[idx]))]
            for _, fmask, fint, _ in component_results:
                row.append(_fmt(float(fint[idx])) if fmask[idx] else "")
            w.writerow(row)

        # 3) Positions
        section("Positions")
        w.writerow(["Peak_ref"] + comp_labels)
        for idx, center in enumerate(peak_centers):
            row = [f"{center:.4f}"]
            for _, fmask, _, fpos in component_results:
                row.append(f"{fpos[idx]:.4f}" if (fmask[idx] and not np.isnan(fpos[idx])) else "")
            w.writerow(row)

        # 4) Presence
        section("Presence")
        w.writerow(["Peak"] + comp_labels)
        for idx, center in enumerate(peak_centers):
            row = [f"{center:.4f}"]
            for _, fmask, _, _ in component_results:
                row.append("1" if fmask[idx] else "0")
            w.writerow(row)

        # 5) Fit data (optional)
        if fit_data is not None:
            x_f, y_f, yerr_f, amps_f, centers_f = fit_data
            y_corr_f = subtract_background_cfg(y_f, cfg)
            _sig = cfg.sigma() if sigmas is None else sigmas
            _gam = cfg.gamma() if gammas is None else gammas
            total_f, comps_f = compute_model(x_f, amps_f, centers_f,
                                             _sig, _gam, cfg.profile)
            section("Fit_Data")
            w.writerow(["x", "Raw_y", "Error", "BG_corrected", "Sum_Fit"]
                       + [f"{c:.1f}" for c in centers_f])
            for k in range(len(x_f)):
                row = [f"{x_f[k]:.6g}", f"{y_f[k]:.6g}", f"{yerr_f[k]:.6g}",
                       f"{y_corr_f[k]:.6g}", f"{total_f[k]:.6g}"]
                row += [f"{c[k]:.6g}" for c in comps_f]
                w.writerow(row)

        # 6) Components_Raw (optional) -- side-by-side (x, y, err) blocks
        if component_results:
            blocks, max_len = [], 0
            for cf, _fm, _fi, _fp in component_results:
                try:
                    xs, ys, es, _ = load_xy(cf, cfg)
                    blocks.append((cf, xs, ys, es))
                    max_len = max(max_len, len(xs))
                except Exception:
                    blocks.append((cf, None, None, None))
            section("Components_Raw")
            top = []
            for cf, _, _, _ in blocks:
                lab = _ascii(os.path.splitext(os.path.basename(cf))[0])
                top.extend([f"{lab}:x", f"{lab}:y", f"{lab}:err"])
            w.writerow(top)
            for k in range(max_len):
                row = []
                for _, xs, ys, es in blocks:
                    if xs is None or k >= len(xs):
                        row.extend(["", "", ""])
                    else:
                        row.extend([f"{xs[k]:.6g}", f"{ys[k]:.6g}", f"{es[k]:.6g}"])
                w.writerow(row)

    print(f"  CSV saved: {out_path}")
    return [out_path]


# ============================================================
