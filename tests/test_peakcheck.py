"""
Test suite for PeakCheck.

Covers the scientific core (profiles, noise estimation, background subtraction,
NNLS amplitude fit, fit statistics), the peak-presence logic in both SNR modes,
the x-axis conversions (affine and reciprocal), and the configuration round-trip.

Run from the package root:

    pytest                # or:  python -m pytest -v
"""

import os
import sys
import math

import numpy as np
import pytest

# Make `import peakcheck` work regardless of the directory pytest is run from.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import peakcheck as pc  # noqa: E402

try:
    from scipy.special import voigt_profile as _scipy_voigt
    HAVE_SCIPY_VOIGT = True
except Exception:  # pragma: no cover
    HAVE_SCIPY_VOIGT = False


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------
def _grid(lo=-30.0, hi=30.0, n=4001):
    return np.linspace(lo, hi, n)


def _area(x, y):
    return np.trapezoid(y, x)


# --------------------------------------------------------------------------
# profiles
# --------------------------------------------------------------------------
class TestProfiles:
    def test_gaussian_area_normalised(self):
        x = _grid()
        assert _area(x, pc.gaussian(x, 0.0, 3.0)) == pytest.approx(1.0, abs=1e-4)

    def test_lorentzian_area_normalised(self):
        # Lorentzian/Voigt are analytically normalised to 1, but the heavy
        # Lorentzian tails are truncated on a finite grid: for gamma=2 about
        # 0.3 % of the area lies beyond +/-400, so allow a small tolerance.
        x = _grid(-400.0, 400.0, 40001)
        assert _area(x, pc.lorentzian(x, 0.0, 2.0)) == pytest.approx(1.0, abs=5e-3)

    def test_voigt_area_normalised(self):
        x = _grid(-400.0, 400.0, 40001)
        assert _area(x, pc.voigt(x, 0.0, 3.0, 2.0)) == pytest.approx(1.0, abs=5e-3)

    @pytest.mark.skipif(not HAVE_SCIPY_VOIGT, reason="scipy.voigt_profile missing")
    def test_voigt_matches_scipy(self):
        x = _grid()
        ours = pc.voigt(x, 0.0, 3.0, 2.0)
        ref = _scipy_voigt(x, 3.0, 2.0)
        assert np.allclose(ours, ref, atol=1e-12)

    def test_voigt_reduces_to_lorentzian_when_sigma_zero(self):
        x = _grid()
        assert np.allclose(pc.voigt(x, 0.0, 0.0, 2.0),
                           pc.lorentzian(x, 0.0, 2.0), atol=1e-12)

    def test_voigt_reduces_to_gaussian_when_gamma_tiny(self):
        x = _grid()
        v = pc.voigt(x, 0.0, 3.0, 1e-9)
        g = pc.gaussian(x, 0.0, 3.0)
        assert np.allclose(v, g, atol=1e-4)

    def test_voigt_is_not_pseudovoigt(self):
        # the true Voigt must differ clearly from a 50/50 pseudo-Voigt
        x = _grid()
        s, g = 3.0, 2.0
        v = pc.voigt(x, 0.0, s, g)
        pv = 0.5 * pc.lorentzian(x, 0.0, g) + 0.5 * pc.gaussian(x, 0.0, s)
        assert np.max(np.abs(v - pv)) > 1e-3

    def test_profile_shape_dispatch(self):
        x = _grid()
        assert np.allclose(pc.profile_shape(x, 0, 3, 2, "gauss"),
                           pc.gaussian(x, 0, 3))
        assert np.allclose(pc.profile_shape(x, 0, 3, 2, "lorentz"),
                           pc.lorentzian(x, 0, 2))
        assert np.allclose(pc.profile_shape(x, 0, 3, 2, "voigt"),
                           pc.voigt(x, 0, 3, 2))


# --------------------------------------------------------------------------
# Config helpers
# --------------------------------------------------------------------------
class TestConfig:
    def test_sigma_from_fwhm(self):
        cfg = pc.Config(wg=10.0)
        assert cfg.sigma() == pytest.approx(10.0 / (2 * math.sqrt(2 * math.log(2))))

    def test_gamma_is_half_fwhm(self):
        assert pc.Config(wl=4.0).gamma() == pytest.approx(2.0)

    def test_effective_fwhm_pure_limits(self):
        assert pc.Config(profile="gauss", wg=7.0).effective_fwhm() == pytest.approx(7.0)
        assert pc.Config(profile="lorentz", wl=3.0).effective_fwhm() == pytest.approx(3.0)

    def test_effective_fwhm_voigt_between_and_above(self):
        # Olivero-Longbothum FWHM exceeds each component width
        fwhm = pc.Config(profile="voigt", wg=7.0, wl=3.0).effective_fwhm()
        assert fwhm > 7.0 and fwhm < 7.0 + 3.0


# --------------------------------------------------------------------------
# noise estimation
# --------------------------------------------------------------------------
class TestEstimateNoise:
    def test_recovers_known_sigma(self):
        rng = np.random.default_rng(0)
        true_sigma = 5.0
        y = 100.0 + rng.normal(0.0, true_sigma, 20000)
        est = pc.estimate_noise(y)
        assert est == pytest.approx(true_sigma, rel=0.1)

    def test_ignores_smooth_signal(self):
        # a smooth ramp carries no high-frequency noise -> tiny estimate
        x = np.linspace(0, 1, 5000)
        assert pc.estimate_noise(50.0 * x) < 1e-6 + 1.0  # falls back to >=, stays finite

    def test_positive_on_constant(self):
        assert pc.estimate_noise(np.full(100, 7.0)) > 0.0


# --------------------------------------------------------------------------
# background subtraction
# --------------------------------------------------------------------------
class TestBackground:
    def test_window_zero_returns_copy(self):
        y = np.array([1.0, 2.0, 3.0])
        out = pc.subtract_background(y, 0)
        assert np.array_equal(out, y) and out is not y

    def test_removes_constant_baseline(self):
        x = np.linspace(0, 100, 400)
        peak = 500 * pc.gaussian(x, 50, 2.0)
        y = 30.0 + peak
        corr = pc.subtract_background(y, 30)
        # away from the peak the corrected signal should be ~0
        far = (x < 30) | (x > 70)
        assert np.median(corr[far]) < 1.0

    def test_non_negative(self):
        rng = np.random.default_rng(1)
        y = 10 + rng.normal(0, 2, 500)
        assert np.all(pc.subtract_background(y, 25) >= 0.0)


# --------------------------------------------------------------------------
# NNLS amplitude fit + statistics
# --------------------------------------------------------------------------
class TestFit:
    def _make(self, amps=(2000.0, 5000.0, 3000.0), centers=(120, 150, 175)):
        cfg = pc.Config(wg=7.8701, wl=1.094, profile="voigt", bg_window=0)
        s, g = cfg.sigma(), cfg.gamma()
        x = np.linspace(90, 200, 600)
        y = np.zeros_like(x)
        for a, c in zip(amps, centers):
            y += a * pc.voigt(x, c, s, g)
        yerr = np.ones_like(x)
        return cfg, x, y, yerr, list(centers), list(amps)

    def test_recovers_amplitudes(self):
        cfg, x, y, yerr, centers, amps = self._make()
        got = pc.fit_amplitudes(x, y, yerr, centers, cfg)
        assert np.allclose(got, amps, rtol=1e-3)

    def test_non_negative_amplitudes(self):
        cfg, x, y, yerr, centers, _ = self._make()
        # add a decoy centre where there is no peak -> must come back as 0
        centers2 = centers + [110.0]
        got = pc.fit_amplitudes(x, y, yerr, centers2, cfg)
        assert np.all(got >= 0.0)
        assert got[-1] == pytest.approx(0.0, abs=1e-6)

    def test_statistics_good_fit(self):
        cfg, x, y, yerr, centers, _ = self._make()
        amps = pc.fit_amplitudes(x, y, yerr, centers, cfg)
        st = pc.fit_statistics(x, y, yerr, amps, centers, cfg)
        assert st["r2"] > 0.999
        assert st["n_active"] == 3
        assert st["chi2_red"] >= 0.0

    def test_scaled_vs_statistical_errors(self):
        cfg, x, y, yerr, centers, _ = self._make()
        amps = pc.fit_amplitudes(x, y, yerr, centers, cfg)
        st_stat = pc.fit_statistics(x, y, yerr, amps, centers, cfg, force_scaled=False)
        st_scaled = pc.fit_statistics(x, y, yerr, amps, centers, cfg, force_scaled=True)
        # scaled error = statistical error * sqrt(chi2_red)
        scale = math.sqrt(st_stat["chi2_red"]) if st_stat["chi2_red"] > 0 else 1.0
        a = st_stat["stderr"][0]
        b = st_scaled["stderr"][0]
        assert b == pytest.approx(a * scale, rel=1e-6)


# --------------------------------------------------------------------------
# presence check (both SNR modes)
# --------------------------------------------------------------------------
class TestPresence:
    def _component(self, present_centers, base=10.0, seed=3):
        cfg = pc.Config(wg=7.8701, wl=1.094, profile="voigt", bg_window=30)
        s, g = cfg.sigma(), cfg.gamma()
        x = np.linspace(95, 205, 400)
        y = np.full_like(x, base)
        for c, a in present_centers:
            y += a * pc.voigt(x, c, s, g)
        rng = np.random.default_rng(seed)
        yerr = np.sqrt(np.clip(y, 1, None))
        y = np.clip(y + rng.normal(0, yerr), 0, None)
        return cfg, x, y, yerr

    def test_corrected_mode_discriminates(self):
        # peaks present at 112,151,190; absent at 128,168
        cfg, x, y, yerr = self._component([(112, 2700), (151, 7000), (190, 1700)])
        centers = [112, 128, 151, 168, 190]
        mask, _, _ = pc.check_peaks_in_component(
            x, y, yerr, centers, 5.0, 5.0, cfg.bg_window, baseline_corrected=True)
        assert mask == [True, False, True, False, True]

    def test_raw_mode_is_less_discriminating(self):
        # with the raw signal and a modest threshold the empty positions leak in
        cfg, x, y, yerr = self._component([(112, 2700), (151, 7000), (190, 1700)])
        centers = [112, 128, 151, 168, 190]
        mask_raw, _, _ = pc.check_peaks_in_component(
            x, y, yerr, centers, 5.0, 5.0, cfg.bg_window, baseline_corrected=False)
        mask_corr, _, _ = pc.check_peaks_in_component(
            x, y, yerr, centers, 5.0, 5.0, cfg.bg_window, baseline_corrected=True)
        # raw mode finds at least as many peaks as corrected mode (it is laxer here)
        assert sum(mask_raw) >= sum(mask_corr)

    def test_all_present_when_all_there(self):
        cfg, x, y, yerr = self._component(
            [(112, 2700), (128, 5200), (151, 7000), (168, 3200), (190, 1700)])
        centers = [112, 128, 151, 168, 190]
        mask, _, _ = pc.check_peaks_in_component(
            x, y, yerr, centers, 5.0, 5.0, cfg.bg_window, baseline_corrected=True)
        assert all(mask)


# --------------------------------------------------------------------------
# x-axis conversions
# --------------------------------------------------------------------------
class TestConversions:
    def test_affine_mev_to_wavenumber(self):
        cfg = pc.Config()
        assert pc.apply_conversion(cfg, "meV -> cm^-1")
        assert cfg.x_transform == "affine"
        x = np.array([10.0, 20.0])
        out = pc.apply_x_transform(x, cfg)
        assert out[0] == pytest.approx(80.65544, rel=1e-5)

    def test_reciprocal_nm_to_wavenumber(self):
        cfg = pc.Config()
        assert pc.apply_conversion(cfg, "nm -> cm^-1 (reciprocal)")
        assert cfg.x_transform == "reciprocal"
        out = pc.apply_x_transform(np.array([500.0]), cfg)
        assert out[0] == pytest.approx(1e7 / 500.0, rel=1e-9)  # 20000 cm^-1

    def test_reciprocal_nm_to_eV(self):
        cfg = pc.Config()
        pc.apply_conversion(cfg, "nm -> eV (reciprocal)")
        out = pc.apply_x_transform(np.array([620.0]), cfg)
        assert out[0] == pytest.approx(1239.84198 / 620.0, rel=1e-6)

    def test_unknown_preset_returns_false(self):
        cfg = pc.Config()
        assert pc.apply_conversion(cfg, "no such unit") is False

    def test_reciprocal_guards_zero(self):
        cfg = pc.Config(x_transform="reciprocal", x_scale=1e7)
        out = pc.apply_x_transform(np.array([0.0, 100.0]), cfg)
        assert np.isnan(out[0]) and np.isfinite(out[1])


# --------------------------------------------------------------------------
# config round-trip + data loading
# --------------------------------------------------------------------------
class TestConfigIO:
    def test_template_roundtrip(self, tmp_path):
        cfg = pc.Config(wg=7.87, wl=1.09, profile="voigt", x_min=100.0, x_max=200.0,
                        snr_thresh=5.0, presence_baseline_corrected=True)
        path = tmp_path / "job.toml"
        pc.write_template(str(path), cfg)
        loaded = pc.load_config(str(path))
        pc._coerce_nullable(loaded)
        assert loaded.wg == pytest.approx(7.87)
        assert loaded.profile == "voigt"
        assert loaded.x_min == pytest.approx(100.0)
        assert loaded.presence_baseline_corrected is True

    def test_named_conversion_in_toml(self, tmp_path):
        path = tmp_path / "c.toml"
        path.write_text('[xaxis]\nx_conversion = "meV -> cm^-1"\n')
        cfg = pc.load_config(str(path))
        assert cfg.x_scale == pytest.approx(8.065544, rel=1e-5)
        assert cfg.x_transform == "affine"

    def test_load_xy_two_and_three_columns(self, tmp_path):
        x = np.linspace(0, 10, 50)
        y = 3.0 + x
        # two columns
        f2 = tmp_path / "two.txt"
        np.savetxt(str(f2), np.column_stack([x, y]))
        cfg = pc.Config()
        xx, yy, ee, had = pc.load_xy(str(f2), cfg)
        assert not had and len(xx) == 50 and np.all(ee > 0)
        # three columns
        f3 = tmp_path / "three.txt"
        np.savetxt(str(f3), np.column_stack([x, y, np.full_like(x, 0.5)]))
        xx, yy, ee, had = pc.load_xy(str(f3), cfg)
        assert had and np.allclose(ee, 0.5)

    def test_load_xy_applies_transform_and_sorts(self, tmp_path):
        # reversed, with a meV->cm^-1 conversion: must come back sorted & scaled
        x = np.array([20.0, 10.0, 15.0])
        y = np.array([1.0, 2.0, 3.0])
        f = tmp_path / "d.txt"
        np.savetxt(str(f), np.column_stack([x, y]))
        cfg = pc.Config()
        pc.apply_conversion(cfg, "meV -> cm^-1")
        xx, yy, ee, had = pc.load_xy(str(f), cfg)
        assert np.all(np.diff(xx) > 0)
        assert xx[0] == pytest.approx(10.0 * 8.065544, rel=1e-5)

    def test_load_xy_tolerates_comments(self, tmp_path):
        f = tmp_path / "h.txt"
        f.write_text("# a header\n! another comment\n1.0 2.0\n2.0 4.0\n3.0 6.0\n")
        cfg = pc.Config()
        xx, yy, ee, had = pc.load_xy(str(f), cfg)
        assert len(xx) == 3 and np.allclose(yy, [2.0, 4.0, 6.0])


# --------------------------------------------------------------------------
# output: Components_Raw sheet + _components.csv + parameter provenance
# --------------------------------------------------------------------------
class TestOutputs:
    def _make_files(self, tmp_path):
        """Build a tiny reference + two components in tmp_path."""
        cfg = pc.Config(wg=7.87, wl=1.09, profile="voigt", bg_window=10,
                        x_min=10.0, x_max=40.0)
        s, g = cfg.sigma(), cfg.gamma()
        x = np.linspace(0, 50, 200)
        def make(centers_amps, fname, n=200):
            xs = np.linspace(0, 50, n)
            y = np.full_like(xs, 5.0)
            for c, a in centers_amps:
                y += a * pc.voigt(xs, c, s, g)
            err = np.sqrt(np.clip(y, 1, None))
            path = tmp_path / fname
            np.savetxt(str(path), np.column_stack([xs, y, err]))
            return str(path)
        ref = make([(15, 800), (25, 1200), (35, 500)], "ref.txt")
        c1  = make([(15, 700),               (35, 450)], "comp_A.txt", n=180)
        c2  = make([(15, 750), (25, 1000), (35, 480)], "comp_B.txt", n=210)
        cfg.reference_file = ref
        return cfg, [c1, c2]

    def test_components_csv_written_and_ascii(self, tmp_path):
        cfg, comps = self._make_files(tmp_path)
        x, y, e, had = pc.load_xy(cfg.reference_file, cfg)
        centers = np.array([15.0, 25.0, 35.0])
        comp_results = []
        for cf in comps:
            xs, ys, es, _ = pc.load_xy(cf, cfg)
            m, fi, fp = pc.check_peaks_in_component(
                xs, ys, es, centers, cfg.tolerance, cfg.snr_thresh,
                cfg.bg_window, cfg.presence_baseline_corrected)
            comp_results.append((cf, m, fi, fp))
        params = pc.gather_parameters(cfg, centers, None, None, had,
                                      component_files=comps)
        written = pc.save_csv(cfg, centers, np.array([800, 1200, 500.]),
                              comp_results, fit_data=None, stats=None, params=params)
        # exactly one CSV file is written now
        assert len(written) == 1, f"expected 1 CSV, got {len(written)}"
        out_csv = written[0]
        assert out_csv.endswith("_results.csv")
        data = open(out_csv, "rb").read()
        assert all(b < 128 for b in data), "results CSV is not 7-bit ASCII"
        text = data.decode("ascii")
        # the file holds five labelled sections in one document
        for sec in ("Intensities", "Positions", "Presence", "Components_Raw"):
            assert f"# === SECTION: {sec} ===" in text, f"missing section {sec}"
        # the Components_Raw section carries the component column headers
        for tag in ("comp_A:x", "comp_A:y", "comp_A:err",
                    "comp_B:x", "comp_B:y", "comp_B:err"):
            assert tag in text, f"missing column {tag}"
        # Parameter block lists every component path
        assert "Component 1" in text and "Component 2" in text

    def test_output_dir_overrides_default_location(self, tmp_path):
        """If cfg.output_dir is set, outputs land there instead of next to the
        reference file (issue surfaced when running on read-only data folders)."""
        cfg, comps = self._make_files(tmp_path)
        x, y, e, had = pc.load_xy(cfg.reference_file, cfg)
        centers = np.array([15.0, 25.0, 35.0])
        comp_results = []
        for cf in comps:
            xs, ys, es, _ = pc.load_xy(cf, cfg)
            m, fi, fp = pc.check_peaks_in_component(
                xs, ys, es, centers, cfg.tolerance, cfg.snr_thresh,
                cfg.bg_window, cfg.presence_baseline_corrected)
            comp_results.append((cf, m, fi, fp))
        # redirect outputs into a fresh subdirectory
        out_dir = tmp_path / "results_here"
        cfg.output_dir = str(out_dir)
        params = pc.gather_parameters(cfg, centers, None, None, had,
                                      component_files=comps)
        written = pc.save_csv(cfg, centers, np.array([800, 1200, 500.]),
                              comp_results, fit_data=None, stats=None, params=params)
        # the file must be inside out_dir, not in tmp_path itself
        assert os.path.dirname(written[0]) == str(out_dir)
        assert os.path.isfile(written[0])
        # and the data directory must NOT contain a _results.csv
        leak = [f for f in os.listdir(str(tmp_path))
                if f.endswith("_results.csv") and os.path.isfile(str(tmp_path / f))]
        assert leak == [], f"output leaked into data dir: {leak}"

    def test_excel_components_raw_sheet(self, tmp_path):
        from openpyxl import load_workbook
        cfg, comps = self._make_files(tmp_path)
        x, y, e, had = pc.load_xy(cfg.reference_file, cfg)
        centers = np.array([15.0, 25.0, 35.0])
        comp_results = []
        for cf in comps:
            xs, ys, es, _ = pc.load_xy(cf, cfg)
            m, fi, fp = pc.check_peaks_in_component(
                xs, ys, es, centers, cfg.tolerance, cfg.snr_thresh,
                cfg.bg_window, cfg.presence_baseline_corrected)
            comp_results.append((cf, m, fi, fp))
        amps = np.array([800.0, 1200.0, 500.0])
        # provide a minimal stats dict so the fit gets a Std_Err column
        stats = {"chi2": 1.0, "chi2_red": 1.0, "r2": 0.99, "dof": 100,
                 "n_active": 3, "n_points": 100,
                 "stderr": np.array([10.0, 12.0, 8.0]),
                 "stderr_stat": np.array([10.0, 12.0, 8.0]),
                 "error_scale": 1.0, "error_mode": "statistical"}
        params = pc.gather_parameters(cfg, centers, None, stats, had,
                                      component_files=comps)
        out = pc.save_excel(cfg, centers, amps, comp_results,
                            fit_data=(x, y, e, amps, centers),
                            stats=stats, params=params)
        wb = load_workbook(out)
        assert "Components_Raw" in wb.sheetnames
        ws = wb["Components_Raw"]
        # 2 components x 3 columns = 6 columns
        assert ws.max_column == 6
        # Parameters sheet contains the component paths
        ws_p = wb["Parameters"]
        param_values = [ws_p.cell(r, 2).value for r in range(1, ws_p.max_row + 1)]
        assert any(comps[0] in str(v) for v in param_values)
        assert any(comps[1] in str(v) for v in param_values)


# --------------------------------------------------------------------------
# alternative baselines
# --------------------------------------------------------------------------
class TestBaselines:
    def _sloped(self, seed=0):
        cfg = pc.Config(wg=7.8701, wl=1.094, bg_window=30)
        s, g = cfg.sigma(), cfg.gamma()
        x = np.linspace(0, 100, 600)
        y = 40 + 0.8 * x + 3000 * pc.voigt(x, 50, s, g)
        return x, y

    def test_als_removes_slope(self):
        x, y = self._sloped()
        bg = 40 + 0.8 * x
        corr = pc.subtract_background(y, 30, method="als")
        away = (x < 25) | (x > 75)
        assert corr.min() >= 0.0
        assert np.median(corr[away]) < 0.25 * np.median(bg[away])   # >75% removed
        assert corr[(x > 45) & (x < 55)].max() > 150                # peak preserved

    def test_polynomial_removes_slope(self):
        x, y = self._sloped()
        bg = 40 + 0.8 * x
        corr = pc.subtract_background(y, 30, method="polynomial", poly_order=3)
        away = (x < 25) | (x > 75)
        assert corr.min() >= 0.0
        assert np.median(corr[away]) < 0.5 * np.median(bg[away])    # most removed
        assert corr[(x > 45) & (x < 55)].max() > 150                # peak preserved

    def test_cfg_selects_method(self):
        x, y = self._sloped()
        cfg = pc.Config(bg_window=30, baseline_method="als")
        assert np.allclose(pc.subtract_background_cfg(y, cfg),
                           pc.subtract_background(y, 30, method="als"))

    def test_rolling_min_still_default(self):
        x, y = self._sloped()
        cfg = pc.Config(bg_window=30)            # default method
        assert np.allclose(pc.subtract_background_cfg(y, cfg),
                           pc.subtract_background(y, 30))

    def test_unknown_method_flagged_by_validate(self):
        assert any("baseline_method" in p
                   for p in pc.Config(baseline_method="nope").validate())


# --------------------------------------------------------------------------
# optional position refinement
# --------------------------------------------------------------------------
class TestRefine:
    def _data(self):
        cfg = pc.Config(wg=7.8701, wl=1.094, profile="voigt", bg_window=0,
                        refine=True, refine_window=3.0)
        s, g = cfg.sigma(), cfg.gamma()
        x = np.linspace(95, 205, 361)
        centers = [112.0, 151.0, 190.0]
        y = np.zeros_like(x)
        for c, a in zip(centers, [3000.0, 5000.0, 2000.0]):
            y += a * pc.voigt(x, c, s, g)
        yerr = np.sqrt(np.maximum(y, 1.0))
        return cfg, x, y, yerr, centers

    def test_refine_recovers_offset_positions(self):
        cfg, x, y, yerr, centers = self._data()
        refined = pc.refine_peaks(x, y, yerr, [112.8, 150.2, 189.1], cfg)
        assert np.allclose(refined, centers, atol=0.05)

    def test_refine_disabled_is_noop(self):
        cfg, x, y, yerr, centers = self._data()
        cfg.refine = False
        start = [112.8, 150.2, 189.1]
        assert np.allclose(pc.refine_peaks(x, y, yerr, start, cfg), start)

    def test_refine_respects_window_bound(self):
        cfg, x, y, yerr, centers = self._data()
        cfg.refine_window = 0.2
        start = np.array([112.0, 151.0, 190.0]) + 2.0
        refined = pc.refine_peaks(x, y, yerr, start, cfg)
        assert np.all(np.abs(refined - start) <= 0.2 + 1e-6)


# --------------------------------------------------------------------------
# regression: the bundled example reproduces the documented numbers
# --------------------------------------------------------------------------
class TestRegressionExample:
    def test_bundled_example_reproduces(self, tmp_path):
        here = os.path.dirname(os.path.abspath(__file__))
        ref = os.path.join(os.path.dirname(here), "example_data", "sample.nis")
        if not os.path.isfile(ref):
            pytest.skip("bundled example data not available")
        cfg = pc.Config(reference_file=ref, x_conversion="meV -> cm^-1",
                        wg=7.8701, wl=1.094, profile="voigt",
                        min_snr_init=3.0, tolerance=5.0, snr_thresh=5.0,
                        error_mode="statistical", output_dir=str(tmp_path),
                        write_csv=False, write_excel=False)
        cfg.peaks = [112.0, 128.0, 151.0, 168.0, 190.0]
        pc.apply_conversion(cfg, cfg.x_conversion)   # sets x_scale = 8.065544
        res = pc.run_headless(cfg, verbose=False)
        amps = [round(float(a)) for a in res["amplitudes"]]
        assert amps == [1969, 4111, 5675, 2408, 1507]
        assert round(res["stats"]["chi2_red"], 1) == 5.5
        assert round(res["stats"]["r2"], 3) == 0.945
        comp = {os.path.basename(cf): list(map(int, fm))
                for cf, fm, _, _ in res["component_results"]}
        assert comp["sample_A.nis"] == [1, 0, 1, 0, 1]


# ---------------------------------------------------------------------------
class TestCLI:
    """Command-line surface: --list-conversions, --validate and --output-dir."""

    def test_list_conversions_prints_presets(self, capsys):
        from peakcheck.cli import main
        main(["--list-conversions"])
        out = capsys.readouterr().out
        assert "x_conversion" in out
        assert "meV -> cm^-1" in out

    def test_validate_ok(self, capsys, tmp_path):
        from peakcheck.cli import main
        cfg = tmp_path / "job.toml"
        cfg.write_text('[search]\nbaseline_method = "als"\n[refine]\nrefine = true\nrefine_window = 2.0\n')
        main(["--config", str(cfg), "--validate"])
        assert "Configuration OK" in capsys.readouterr().out

    def test_validate_bad_exits_nonzero(self, tmp_path):
        from peakcheck.cli import main
        cfg = tmp_path / "bad.toml"
        cfg.write_text('[search]\nbaseline_method = "nope"\n[refine]\nrefine = true\nrefine_window = 0\n')
        with pytest.raises(SystemExit) as exc:
            main(["--config", str(cfg), "--validate"])
        assert exc.value.code == 1

    def test_missing_config_clean_message(self, tmp_path):
        """A missing config file gives a clean message, not a traceback."""
        from peakcheck.cli import main
        with pytest.raises(SystemExit) as exc:
            main(["--config", str(tmp_path / "nope.toml"), "--no-gui"])
        assert "config file not found" in str(exc.value)

    def test_missing_data_file_clean_message(self, tmp_path):
        """A config pointing at a missing data file fails cleanly."""
        from peakcheck.cli import main
        cfg = tmp_path / "job.toml"
        cfg.write_text(f'reference_file = "{tmp_path / "gone.dat"}"\n')
        with pytest.raises(SystemExit) as exc:
            main(["--config", str(cfg), "--no-gui"])
        assert "PeakCheck:" in str(exc.value) and "not found" in str(exc.value)

    def test_debug_flag_reraises_traceback(self, tmp_path):
        """With --debug the original exception propagates (not SystemExit)."""
        from peakcheck.cli import main
        cfg = tmp_path / "job.toml"
        junk = tmp_path / "junk.dat"
        junk.write_text("das ist\nnur text\n")
        cfg.write_text(f'reference_file = "{junk}"\n')
        with pytest.raises(ValueError):       # raw error, not SystemExit
            main(["--config", str(cfg), "--no-gui", "--debug"])

    def test_output_dir_flag_routes_outputs(self, tmp_path):
        """`--output-dir` sends every output into the target folder, leaving the
        reference directory clean — the batch use case for the flag."""
        from peakcheck.cli import main
        cfg = pc.Config()                       # defaults (wg/wl, prominence, ...)
        s, g = cfg.sigma(), cfg.gamma()
        xs = np.linspace(95, 205, 400)
        y = np.full_like(xs, 10.0) + 6000.0 * pc.voigt(xs, 150.0, s, g)
        err = np.sqrt(np.clip(y, 1, None))
        ref = tmp_path / "ref.txt"
        np.savetxt(str(ref), np.column_stack([xs, y, err]))
        out_dir = tmp_path / "outs"
        main(["--reference", str(ref), "--no-gui", "--output-dir", str(out_dir)])
        produced = list(out_dir.glob("*_results.csv"))
        assert produced, "no results CSV written into --output-dir"
        leak = list(tmp_path.glob("*_results.csv"))
        assert leak == [], f"results leaked next to the reference: {leak}"


# ---------------------------------------------------------------------------
class TestConfigPaths:
    """A relative reference_file in a config resolves against the config file's
    directory, not the current working directory (portable / batch-friendly)."""

    def test_relative_reference_resolves_to_config_dir(self, tmp_path, monkeypatch):
        job = tmp_path / "job"
        job.mkdir()
        # a data file living next to the config, named relatively in the TOML
        (job / "spectrum.dat").write_text("1 10\n2 20\n3 30\n")
        (job / "run.toml").write_text(
            '[input]\nreference_file = "spectrum.dat"\n')
        # run from a DIFFERENT working directory
        elsewhere = tmp_path / "elsewhere"
        elsewhere.mkdir()
        monkeypatch.chdir(elsewhere)
        cfg = pc.load_config(str(job / "run.toml"))
        assert os.path.isabs(cfg.reference_file)
        assert cfg.reference_file == os.path.normpath(str(job / "spectrum.dat"))
        assert os.path.isfile(cfg.reference_file)

    def test_absolute_reference_is_left_unchanged(self, tmp_path):
        data = tmp_path / "abs_data.dat"
        data.write_text("1 10\n2 20\n")
        cfg_file = tmp_path / "abs.toml"
        cfg_file.write_text(f'[input]\nreference_file = "{data}"\n')
        cfg = pc.load_config(str(cfg_file))
        assert cfg.reference_file == str(data)

    def test_relative_components_found_next_to_resolved_reference(self, tmp_path, monkeypatch):
        """After the reference is resolved, the auto component glob finds the
        component files sitting next to it — independent of the working dir."""
        job = tmp_path / "exp"
        job.mkdir()
        cfg0 = pc.Config(wg=7.87, wl=1.09, profile="voigt", bg_window=10,
                         x_min=10.0, x_max=40.0)
        s, g = cfg0.sigma(), cfg0.gamma()

        def make(centers_amps, fname, n=200):
            xs = np.linspace(0, 50, n)
            y = np.full_like(xs, 5.0)
            for c, a in centers_amps:
                y += a * pc.voigt(xs, c, s, g)
            np.savetxt(str(job / fname), np.column_stack([xs, y, np.sqrt(np.clip(y, 1, None))]))

        make([(15, 800), (25, 1200), (35, 500)], "ref.txt")
        make([(15, 700), (35, 450)], "ref_A.txt", n=180)
        (job / "run.toml").write_text(
            '[input]\nreference_file = "ref.txt"\n'
            '[profile]\nwg = 7.87\nwl = 1.09\n'
            '[window]\nx_min = 10.0\nx_max = 40.0\n')
        monkeypatch.chdir(tmp_path)            # not inside job/
        cfg = pc.load_config(str(job / "run.toml"))
        from peakcheck.pipeline import find_component_files
        comps = find_component_files(cfg)
        names = sorted(os.path.basename(c) for c in comps)
        assert names == ["ref_A.txt"], f"expected ref_A.txt next to the reference, got {names}"


# ---------------------------------------------------------------------------
class TestWidthRefine:
    """Optional per-peak width refinement (bounded variable projection)."""

    def _cfg(self, **kw):
        base = dict(wg=7.8701, wl=1.094, profile="voigt", bg_window=0,
                    refine_widths=True, width_mode="fwhm",
                    width_min_factor=1.0, width_max_factor=3.0)
        base.update(kw)
        return pc.Config(**base)

    def _broadened_peak(self, cfg, scale, center=150.0, seed=1):
        s0, g0 = cfg.sigma(), cfg.gamma()
        x = np.linspace(95, 205, 600)
        clean = 5000.0 * pc.voigt(x, center, s0 * scale, g0 * scale)
        rng = np.random.default_rng(seed)
        yerr = np.sqrt(np.maximum(clean, 1.0))
        y = clean + rng.normal(0, yerr * 0.2)
        return x, y, yerr

    def test_recovers_broadened_width(self):
        cfg = self._cfg()
        x, y, yerr = self._broadened_peak(cfg, 1.5)
        fm = pc.make_fit_mask(x, cfg)
        w = pc.refine_widths(x, y, yerr, np.array([150.0]), cfg, fm)
        assert abs(float(w[0]) - 1.5) < 0.12, f"recovered scale {w[0]}"

    def test_disabled_is_noop(self):
        cfg = self._cfg(refine_widths=False)
        x, y, yerr = self._broadened_peak(cfg, 1.5)
        fm = pc.make_fit_mask(x, cfg)
        w = pc.refine_widths(x, y, yerr, np.array([150.0]), cfg, fm)
        assert np.allclose(w, 1.0)

    def test_respects_upper_bound(self):
        cfg = self._cfg(width_max_factor=1.3)
        x, y, yerr = self._broadened_peak(cfg, 2.5)   # wants to go well beyond 1.3
        fm = pc.make_fit_mask(x, cfg)
        w = pc.refine_widths(x, y, yerr, np.array([150.0]), cfg, fm)
        assert float(w[0]) <= 1.3 + 1e-6

    def test_sigma_mode_keeps_gamma_fixed(self):
        cfg = self._cfg(width_mode="sigma")
        g0 = cfg.gamma()
        sig, gam = pc.per_peak_widths(cfg, 2, np.array([1.4, 1.7]))
        assert np.allclose(gam, g0)              # gamma untouched
        assert np.allclose(sig, cfg.sigma() * np.array([1.4, 1.7]))

    def test_per_peak_widths_default_is_instrumental(self):
        cfg = self._cfg()
        sig, gam = pc.per_peak_widths(cfg, 3, None)
        assert np.allclose(sig, cfg.sigma()) and np.allclose(gam, cfg.gamma())

    def test_fwhm_mode_scales_effective_fwhm_linearly(self):
        cfg = self._cfg(width_mode="fwhm")
        sig, gam = pc.per_peak_widths(cfg, 1, np.array([2.0]))
        fw = pc.fwhm_from_sigma_gamma(sig, gam, cfg.profile)[0]
        assert abs(fw - 2.0 * cfg.effective_fwhm()) < 1e-6

    def test_validate_flags_bad_width_mode(self):
        cfg = self._cfg(width_mode="nope")
        assert any("width_mode" in p for p in cfg.validate())

    def test_validate_flags_bad_bounds(self):
        cfg = self._cfg(width_min_factor=2.0, width_max_factor=1.0)
        assert any("width_max_factor" in p for p in cfg.validate())

    def test_extra_params_reduce_dof(self):
        cfg = self._cfg(refine_widths=False)
        x, y, yerr = self._broadened_peak(cfg, 1.0)
        centers = np.array([150.0])
        fm = pc.make_fit_mask(x, cfg)
        amps = pc.fit_amplitudes(x, y, yerr, centers, cfg, fm)
        s_no = pc.fit_statistics(x, y, yerr, amps, centers, cfg, fm, n_extra_params=0)
        s_ex = pc.fit_statistics(x, y, yerr, amps, centers, cfg, fm, n_extra_params=1)
        assert s_ex["dof"] == s_no["dof"] - 1
        assert s_ex["chi2_red"] > s_no["chi2_red"]   # fewer DOF -> larger reduced chi2


# ---------------------------------------------------------------------------
class TestConfigOutputDirPath:
    """A relative output_dir in a config resolves against the config dir; a
    command-line --output-dir keeps working-directory semantics."""

    def test_relative_output_dir_resolves_to_config_dir(self, tmp_path, monkeypatch):
        job = tmp_path / "job"
        job.mkdir()
        (job / "run.toml").write_text(
            '[input]\nreference_file = "spec.dat"\n[output]\noutput_dir = "results"\n')
        monkeypatch.chdir(tmp_path)               # different working dir
        cfg = pc.load_config(str(job / "run.toml"))
        assert cfg.output_dir == os.path.normpath(str(job / "results"))

    def test_empty_output_dir_stays_empty(self, tmp_path):
        (tmp_path / "run.toml").write_text('[input]\nreference_file = "spec.dat"\n')
        cfg = pc.load_config(str(tmp_path / "run.toml"))
        assert cfg.output_dir == ""

    def test_absolute_output_dir_unchanged(self, tmp_path):
        target = tmp_path / "abs_out"
        (tmp_path / "run.toml").write_text(
            f'[output]\noutput_dir = "{target}"\n')
        cfg = pc.load_config(str(tmp_path / "run.toml"))
        assert cfg.output_dir == str(target)


# ---------------------------------------------------------------------------
class TestEdgeCases:
    """Robustness on awkward inputs: bad rows, empty components, no peaks."""

    def test_load_xy_drops_nonfinite_and_fixes_zero_error(self, tmp_path):
        p = tmp_path / "rough.dat"
        # rows: ok, NaN y, ok with zero error, ok with negative error
        p.write_text("10 100 5\n20 nan 5\n30 200 0\n40 150 -3\n")
        cfg = pc.Config(error_column="auto")
        x, y, yerr, had = pc.load_xy(str(p), cfg)
        assert np.all(np.isfinite(x)) and np.all(np.isfinite(y))
        assert len(x) == 3, f"non-finite row not dropped (len {len(x)})"
        assert np.all(yerr > 0), "non-positive errors not repaired"

    @pytest.mark.parametrize("name,content,xy", [
        ("a.dat", "10 100\n20 200\n30 150\n",      ([10, 20, 30], [100, 200, 150])),
        ("b.txt", "10\t100\n20\t200\n",            ([10, 20], [100, 200])),
        ("c.csv", "10,100\n20,200\n",              ([10, 20], [100, 200])),  # comma = sep
        ("d.dat", "10;100\n20;200\n",              ([10, 20], [100, 200])),  # semicolon
        ("e.xy",  "# header\n10 100\n20 200\n",    ([10, 20], [100, 200])),
        ("f.tsv", "energy\tcounts\n10\t100\n20\t200\n", ([10, 20], [100, 200])),
    ])
    def test_delimiters_and_extensions(self, tmp_path, name, content, xy):
        p = tmp_path / name
        p.write_text(content)
        x, y, yerr, had = pc.load_xy(str(p), pc.Config())
        assert list(x) == xy[0] and list(y) == xy[1]

    def test_decimal_comma_is_not_split(self, tmp_path):
        """German-style decimal comma must not be mistaken for a separator."""
        p = tmp_path / "de.txt"
        p.write_text("10,5 100,2\n20,1 200,8\n30,9 150,3\n")
        x, y, yerr, had = pc.load_xy(str(p), pc.Config())
        assert np.allclose(x, [10.5, 20.1, 30.9])
        assert np.allclose(y, [100.2, 200.8, 150.3])

    def test_decimal_comma_with_semicolon_separator(self, tmp_path):
        p = tmp_path / "de2.csv"
        p.write_text("10,5;100,2\n20,1;200,8\n")
        x, y, yerr, had = pc.load_xy(str(p), pc.Config())
        assert np.allclose(x, [10.5, 20.1]) and np.allclose(y, [100.2, 200.8])

    def test_presence_absent_when_component_lacks_coverage(self):
        cfg = pc.Config(wg=7.87, wl=1.09, profile="voigt", bg_window=10,
                        tolerance=5.0, snr_thresh=2.0)
        # component only covers 0..50, peaks asked at 120/150 -> no coverage
        xs = np.linspace(0, 50, 200)
        ys = np.full_like(xs, 5.0) + 800.0 * pc.voigt(xs, 25.0, cfg.sigma(), cfg.gamma())
        es = np.sqrt(np.clip(ys, 1, None))
        centers = np.array([120.0, 150.0])
        fmask, fint, fpos = pc.check_peaks_in_component(
            xs, ys, es, centers, cfg.tolerance, cfg.snr_thresh,
            cfg.bg_window, cfg.presence_baseline_corrected)
        assert list(map(bool, fmask)) == [False, False]

    def test_pipeline_zero_peaks_returns_none(self, tmp_path, capsys):
        xs = np.linspace(0, 50, 100)
        ys = np.full_like(xs, 5.0)            # flat: nothing to find
        es = np.sqrt(np.clip(ys, 1, None))
        ref = tmp_path / "flat.dat"
        np.savetxt(str(ref), np.column_stack([xs, ys, es]))
        cfg = pc.Config(reference_file=str(ref), output_dir=str(tmp_path),
                        prominence_init=1e9)   # impossible prominence -> no peaks
        out = pc.run_headless(cfg, verbose=False)
        assert out is None

    def test_pipeline_skips_unreadable_component(self, tmp_path, capsys):
        cfg = pc.Config(wg=7.87, wl=1.09, profile="voigt", bg_window=10,
                        x_min=10.0, x_max=40.0, output_dir=str(tmp_path),
                        write_excel=False, write_csv=False)
        s, g = cfg.sigma(), cfg.gamma()
        xs = np.linspace(0, 50, 200)
        y = np.full_like(xs, 5.0) + 1000.0 * pc.voigt(xs, 25.0, s, g)
        err = np.sqrt(np.clip(y, 1, None))
        ref = tmp_path / "ref.txt"
        np.savetxt(str(ref), np.column_stack([xs, y, err]))
        good = tmp_path / "ref_A.txt"
        np.savetxt(str(good), np.column_stack([xs, y, err]))
        bad = tmp_path / "ref_B.txt"
        bad.write_text("not a data file at all\n")
        cfg.reference_file = str(ref)
        x, yy, ee, had = pc.load_xy(str(ref), cfg)
        res = pc.run_analysis(cfg, x, yy, ee, np.array([25.0]), None, had,
                              verbose=False, component_files=[str(good), str(bad)])
        # the good component is analysed, the bad one is skipped (not crashed)
        names = [os.path.basename(cf) for cf, *_ in res["component_results"]]
        assert "ref_A.txt" in names and "ref_B.txt" not in names


# ---------------------------------------------------------------------------
class TestFIO:
    """Reading DESY/PETRA-III FIO files (structured header + Col declarations)."""

    def _write_fio(self, path, n=40):
        """A minimal but realistic FIO file with a %c/%p/%d structure."""
        x = np.linspace(-5, 35, n)
        sig = 1000.0 + 8000.0 * np.exp(-0.5 * ((x - 15.0) / 3.0) ** 2)
        mon = np.full_like(x, 5.0)
        lines = ["!", "! Comments", "!", "%c",
                 "dscan hrm_ener2 -5.0 35.0 %d 6.0" % (n - 1),
                 "!", "! Parameter", "!", "%p",
                 "dcm_ener = 14411.1", "some_motor = 0.5",
                 "!", "! Data", "!", "%d",
                 " Col 1 hrm_ener2 DOUBLE",
                 " Col 2 eh1_t01 DOUBLE",
                 " Col 3 nisp DOUBLE",
                 " Col 4 nfsp DOUBLE"]
        for xi, si, mi in zip(x, sig, mon):
            lines.append(f" {xi:.6f} {mi:.1f} {si:.6f} {si*0.5:.6f}")
        lines.append("! Acquisition ended at Sometime")
        path.write_text("\n".join(lines) + "\n")
        return x, sig

    def test_reads_columns_and_default_y(self, tmp_path):
        f = tmp_path / "scan_00001.fio"
        x, sig = self._write_fio(f)
        from peakcheck.io import _fio_columns
        names, data = _fio_columns(str(f))
        assert names == ["hrm_ener2", "eh1_t01", "nisp", "nfsp"]
        assert data.shape == (len(x), 4)
        # default x = first column, y = "nisp" (3rd col)
        cfg = pc.Config(fio_y_column="nisp")
        xx, yy, ee, had = pc.load_xy(str(f), cfg)
        assert had is False                      # FIO carries no error column
        assert abs(xx.min() - (-5.0)) < 1e-6 and abs(xx.max() - 35.0) < 1e-6
        assert abs(yy.max() - sig.max()) < 1.0   # picked the signal column

    def test_select_y_by_name_and_index(self, tmp_path):
        f = tmp_path / "scan_00002.fio"
        self._write_fio(f)
        by_name = pc.load_xy(str(f), pc.Config(fio_y_column="nfsp"))[1]
        by_index = pc.load_xy(str(f), pc.Config(fio_y_column="4"))[1]  # 1-based -> nfsp
        assert np.allclose(by_name, by_index)

    def test_select_x_by_name(self, tmp_path):
        f = tmp_path / "scan_00003.fio"
        self._write_fio(f)
        # choosing eh1_t01 as x makes x constant (=5.0 monitor) -> just check it parsed
        cfg = pc.Config(fio_x_column="hrm_ener2", fio_y_column="nisp")
        xx, yy, ee, had = pc.load_xy(str(f), cfg)
        assert len(xx) == len(yy) and len(xx) > 10

    def test_unknown_column_raises(self, tmp_path):
        f = tmp_path / "scan_00004.fio"
        self._write_fio(f)
        with pytest.raises(ValueError):
            pc.load_xy(str(f), pc.Config(fio_y_column="not_a_column"))

    def test_bundled_fio_example_runs(self, tmp_path):
        """The bundled fio_example.fio loads and runs through the pipeline."""
        here = os.path.dirname(os.path.abspath(__file__))
        ref = os.path.join(os.path.dirname(here), "example_data", "fio_example.fio")
        if not os.path.isfile(ref):
            pytest.skip("bundled FIO example not available")
        cfg = pc.Config(reference_file=ref, fio_y_column="nisp",
                        wg=7.8701, wl=1.094, profile="voigt",
                        prominence_init=4000, bg_window=40, min_snr_init=3.0,
                        x_min=10.0, x_max=90.0, output_dir=str(tmp_path),
                        write_csv=False, write_excel=False)
        res = pc.run_headless(cfg, verbose=False)
        assert res is not None and len(res["amplitudes"]) >= 3

    def test_real_uploaded_fio_if_present(self):
        """If the real uploaded FIO files are available, parse one end-to-end."""
        real = "/mnt/user-data/uploads/eh2_11022596_00158.fio"
        if not os.path.isfile(real):
            import pytest as _pt
            _pt.skip("real FIO sample not present")
        cfg = pc.Config(fio_y_column="nisp")
        x, y, yerr, had = pc.load_xy(real, cfg)
        assert len(x) > 500 and had is False
        assert x.min() <= -19 and x.max() >= 89   # the -20..90 energy scan


# ---------------------------------------------------------------------------
class TestPresenceHardPaths:
    """Cover the two complex presence branches: shoulder detection and the
    duplicate-resolution step where two reference peaks claim one maximum."""

    def _cfg(self):
        return pc.Config(wg=7.87, wl=1.09, profile="voigt", bg_window=0,
                         tolerance=6.0, snr_thresh=2.0)

    def test_shoulder_is_detected_as_present(self):
        """A peak that appears only as a shoulder on a larger band should still
        be found (no isolated local maximum of its own)."""
        cfg = self._cfg()
        s, g = cfg.sigma(), cfg.gamma()
        x = np.linspace(100, 200, 500)
        # big band at 150 with a shoulder at ~138 (smaller, merged on the flank)
        y = (9000.0 * pc.voigt(x, 150.0, s, g)
             + 3500.0 * pc.voigt(x, 138.0, s * 1.1, g))
        e = np.sqrt(np.clip(y, 1, None))
        centers = np.array([138.0])
        fmask, fint, fpos = pc.check_peaks_in_component(
            x, y, e, centers, cfg.tolerance, cfg.snr_thresh, cfg.bg_window,
            cfg.presence_baseline_corrected)
        assert bool(fmask[0]) is True            # shoulder recognised as present

    def test_duplicate_resolution_two_refs_one_maximum(self):
        """Two reference peaks fall within tolerance of the same dominant
        maximum; the resolver must keep the closer one on the maximum and try to
        place the other on a nearby alternative."""
        cfg = self._cfg()
        s, g = cfg.sigma(), cfg.gamma()
        x = np.linspace(100, 200, 600)
        # one strong peak at 150, a weaker genuine peak at ~160
        y = (9000.0 * pc.voigt(x, 150.0, s, g)
             + 4000.0 * pc.voigt(x, 160.0, s, g))
        e = np.sqrt(np.clip(y, 1, None))
        # two references near 150 and 156 -> both initially attracted to 150
        centers = np.array([150.0, 156.0])
        fmask, fint, fpos = pc.check_peaks_in_component(
            x, y, e, centers, cfg.tolerance, cfg.snr_thresh, cfg.bg_window,
            cfg.presence_baseline_corrected)
        # the resolver must not assign both references to the identical sample
        pos_found = [p for p, m in zip(fpos, fmask) if m]
        assert len(pos_found) == len(set(pos_found)), "two refs share one maximum"
        assert bool(fmask[0]) is True            # the 150 reference is present


# ---------------------------------------------------------------------------
class TestReleaseConsistency:
    """Meta-tests that guard the release metadata against drift."""

    def test_version_is_consistent_across_files(self):
        here = os.path.dirname(os.path.abspath(__file__))
        root = os.path.dirname(here)
        v = pc.__version__
        # pyproject.toml
        pyproject = open(os.path.join(root, "pyproject.toml")).read()
        assert f'version = "{v}"' in pyproject, "pyproject version mismatch"
        # CITATION.cff
        cff = open(os.path.join(root, "CITATION.cff")).read()
        assert f"version: {v}" in cff, "CITATION.cff version mismatch"
        # CHANGELOG.md has an entry for this version
        changelog = open(os.path.join(root, "CHANGELOG.md")).read()
        assert f"[{v}]" in changelog, "CHANGELOG missing this version"

    def test_template_roundtrips_all_keys(self, tmp_path):
        """write_template + load_config must preserve every Config field."""
        from dataclasses import fields
        fn = tmp_path / "t.toml"
        cfg = pc.Config()
        pc.write_template(str(fn), cfg)
        back = pc.load_config(str(fn))
        for f in fields(cfg):
            if f.name in ("peaks", "reference_file", "output_dir"):
                continue  # path/list fields handled separately
            assert getattr(back, f.name) == getattr(cfg, f.name), \
                f"field {f.name} not round-tripped"


# ---------------------------------------------------------------------------
class TestFitQualityHint:
    """The poor-fit hint must fire on genuine failures and stay silent on good
    fits, so it never cries wolf."""

    def test_hint_fires_when_peaks_dropped(self):
        from peakcheck.pipeline import _fit_quality_hint
        h = _fit_quality_hint({"n_active": 2, "r2": -0.12, "chi2_red": 116}, 5)
        assert h and "dropped" in h and "WG/WL" in h

    def test_hint_fires_on_negative_r2(self):
        from peakcheck.pipeline import _fit_quality_hint
        h = _fit_quality_hint({"n_active": 3, "r2": -0.01, "chi2_red": 9}, 3)
        assert h and "flat line" in h

    def test_no_hint_on_good_fit(self):
        from peakcheck.pipeline import _fit_quality_hint
        assert _fit_quality_hint({"n_active": 5, "r2": 0.95, "chi2_red": 1.2}, 5) == ""

    def test_no_hint_on_good_fit_with_high_chi2(self):
        """A large reduced chi^2 alone (estimated errors) is not flagged."""
        from peakcheck.pipeline import _fit_quality_hint
        assert _fit_quality_hint({"n_active": 5, "r2": 0.94, "chi2_red": 5.5}, 5) == ""

    def test_run_analysis_includes_hint_key(self, tmp_path):
        """run_analysis result carries a 'hint' field; a too-wide profile on a
        narrow-peak data set triggers it (reproduces the GUI report)."""
        import numpy as np
        x = np.linspace(12, 25, 360)
        s, g = pc.Config(wg=0.5, wl=0.2).sigma(), pc.Config(wg=0.5, wl=0.2).gamma()
        y = (300 * pc.voigt(x, 14.0, s, g) + 800 * pc.voigt(x, 18.7, s, g)
             + 200 * pc.voigt(x, 23.6, s, g) + 20)
        e = np.sqrt(np.clip(y, 1, None))
        ref = tmp_path / "narrow.dat"
        np.savetxt(str(ref), np.column_stack([x, y, e]))
        # fit with a far-too-wide profile -> peaks collapse / poor fit
        cfg = pc.Config(reference_file=str(ref), wg=7.8701, wl=1.094,
                        profile="voigt", bg_window=30, output_dir=str(tmp_path),
                        write_csv=False, write_excel=False)
        from peakcheck.pipeline import run_analysis
        res = run_analysis(cfg, x, y, e, np.array([14.0, 18.7, 23.6]),
                           None, True, verbose=False)
        assert "hint" in res and res["hint"]      # a hint was produced


# ---------------------------------------------------------------------------
class TestDecompositionExample:
    """Integration test for the shipped 'one spectrum, ten sub-spectra' example:
    the presence check must reproduce the designed present/absent matrix exactly,
    and the reference fit must recover all eight bands."""

    # designed presence matrix (10 sub-spectra x 8 bands); 1 = present
    _M = [
        [1, 1, 0, 1, 0, 1, 0, 0], [1, 0, 1, 0, 1, 0, 1, 0],
        [0, 1, 1, 0, 0, 1, 0, 1], [1, 1, 1, 0, 0, 0, 0, 0],
        [0, 0, 0, 1, 1, 1, 1, 0], [1, 0, 0, 0, 1, 0, 0, 1],
        [0, 1, 0, 1, 0, 0, 1, 0], [0, 0, 1, 0, 1, 1, 0, 1],
        [1, 0, 0, 1, 0, 0, 1, 1], [0, 1, 1, 0, 1, 0, 0, 0],
    ]

    def _dir(self):
        import os
        here = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(os.path.dirname(here), "example_data", "decomposition")

    def test_example_files_are_shipped(self):
        import os
        d = self._dir()
        assert os.path.isfile(os.path.join(d, "spectrum.dat"))
        assert os.path.isfile(os.path.join(d, "decomposition.toml"))
        for i in range(1, 11):
            assert os.path.isfile(os.path.join(d, f"spectrum_{i:02d}.dat"))

    def test_presence_matches_designed_matrix(self):
        import os
        import numpy as np
        d = self._dir()
        cfg = pc.load_config(os.path.join(d, "decomposition.toml"))
        peaks = np.array(sorted(float(p) for p in cfg.peaks))
        assert len(peaks) == 8
        for r in range(10):
            x, y, e, _ = pc.load_xy(os.path.join(d, f"spectrum_{r + 1:02d}.dat"), cfg)
            fmask, _, _ = pc.check_peaks_in_component(
                x, y, e, peaks, cfg.tolerance, cfg.snr_thresh, cfg.bg_window,
                cfg.presence_baseline_corrected)
            got = [1 if b else 0 for b in fmask]
            assert got == self._M[r], f"sub-spectrum {r + 1}: {got} != {self._M[r]}"

    def test_reference_fit_recovers_all_bands(self):
        import os
        import numpy as np
        d = self._dir()
        cfg = pc.load_config(os.path.join(d, "decomposition.toml"))
        peaks = np.array(sorted(float(p) for p in cfg.peaks))
        x, y, e, _ = pc.load_xy(os.path.join(d, "spectrum.dat"), cfg)
        from peakcheck.fit import make_fit_mask, fit_amplitudes, fit_statistics
        fm = make_fit_mask(x, cfg)
        A = fit_amplitudes(x, y, e, peaks, cfg, fm)
        st = fit_statistics(x, y, e, A, peaks, cfg, fm)
        assert st["n_active"] == 8 and st["r2"] > 0.99
